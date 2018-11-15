# -*-coding: utf-8 -*-

import requests
import openpyxl
from bs4 import BeautifulSoup


def get_response(url):
    r"""Get the response of the URL

    :param url: the url of the response
    :type url: str
    :return: :class:`Response <Response>` object
    :rtype: requests.Response
    """

    try:
        headers = {
            'User-Agent':
                'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181\
                 Safari/537.36'
        }
        response = requests.get(url, headers=headers)
        # Change to the appropriate encoding utf-8
        response.encoding = 'utf-8'
        # Check whether the request for this web page succeeded
        if response.status_code == requests.codes.ok:
            print(f"Request of {url} succeeded")
            return response
        else:
            print(f"Request of {url} failed")
            return None
    except requests.RequestException:
        print(f"Request of {url} failed")
        return None


def parse_homepage(response):
    r"""Parse the form of the homepage

    :param response: :class:`Response <response>` object
    :type response: requests.Response
    :return: a generator that generates a dictionary containing a school's name and its corresponding date and url
    :rtype: generator
    """

    # Create a BeautifulSoup object using the parser html5lib
    soup = BeautifulSoup(response.text, features='html5lib')
    # Perform a CSS selection on the BeautifulSoup element to get all the cells of the form
    cells = soup.select('.willnum-body > table > tbody > tr > td')
    # Loop through every row in the form
    print("Retrieving data from the sheet")
    for i in range(3, len(cells), 3):
        try:
            # Get the text of the first column, the university's name
            name = cells[i].string
            # Get the text of the second column, the university's application date
            date = cells[i + 1].string
            # Get the reference of the text of the third column, the university's admission info's link
            link = cells[i + 2].find('a')['href']

            print(name, date, link)

            yield {
                'name': name,
                'date': date,
                'link': link
            }
        # Handle the TypeError exception that occurs when no link or date is provided
        except TypeError:
            print(f"No date, link or both of the school {name} provided")
            # Not yield a dictionary of the school's information
            pass


def write_to_excel(generator, fname):
    r"""Write the information to an excel form

    :param generator: a generator that generates a dictionary containing the school's name and its corresponding date\
     and url
    :type generator: generator
    :param fname: the name of the created excel file
    :type fname: str
    """

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '报考简章'
    sheet.cell(row=1, column=1).value = '高校名单'
    sheet.cell(row=1, column=2).value = '报名时间'
    sheet.cell(row=1, column=3).value = '招生简章'
    row = 2
    for school in generator:
        sheet.cell(row=row, column=1).value = school['name']
        sheet.cell(row=row, column=2).value = school['date']
        sheet.cell(row=row, column=3).value = school['link']
        row += 1

    wb.save(f'{fname}.xlsx')
    print(f"{fname}.xlsx saved")


if __name__ == '__main__':
    homepage_url = 'http://www.eol.cn/html/g/zjswyt/'
    # Download
    homepage_response = get_response(homepage_url)
    school_generator = parse_homepage(homepage_response)
    write_to_excel(school_generator, "浙江省2019年三位一体招生信息")

