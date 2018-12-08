# -*-coding: utf-8 -*-
__author__ = 'Fan Zhang'
__project__ = '三位一体招生信息爬虫'

import requests
import openpyxl
from bs4 import BeautifulSoup
from xml import etree

# a global variable to store the name of schools
school_names = []
# a global variable to store the workbook to record the school info
wb = openpyxl.Workbook()
sheet = wb.active


def get_html(url):
    """
    Get the html content of a web page with the passed url
    :param url: the url of a web page
    :type url: str
    :return: : the html content of the web page
    :rtype: str
    """
    try:
        headers = {
            'User-Agent':
                'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181\
                 Safari/537.36'
        }
        response = requests.get(url, headers=headers)
        # Change to the appropriate encoding utf-8
        response.encoding = 'utf-8'
        # Check whether the request for this web page succeeded
        if response.status_code == requests.codes.ok:
            print(f"Request {url}. Done.")
            return response.text
        else:
            print(f"Request{url}. Failed")
            return None
    except requests.RequestException:
        print(f"Request {url}. Failed")
        return None


def get_homepage_html():
    """
    Get the html content of the homepage (http://www.eol.cn/html/g/zjswyt/)
    :return: the html content of the homepage
    :rtype: str
    """
    homepage_url = 'http://www.eol.cn/html/g/zjswyt/'
    return get_html(homepage_url)


def parse_homepage(html):
    """
    Parse the form of the homepage
    :param html: the html content of the homepage
    :type: str
    :return: a generator that generates a dictionary containing a school's name and its corresponding date and url
    :rtype: generator
    """
    # Create a BeautifulSoup object using the parser html5lib
    soup = BeautifulSoup(html, features='html5lib')
    # Perform a CSS selection on the BeautifulSoup element to get all the cells of the form
    cells = soup.select('.willnum-body > table > tbody > tr > td')
    print("Retrieving data from the sheet")
    # Loop through every row in the form
    for i in range(3, len(cells), 3):
        try:
            # Get the text of the first column, the university's name
            name = cells[i].string
            if name != '中国美术学院':
                # Get the text of the second column, the university's application date
                date = cells[i + 1].string
                # Get the reference of the text of the third column, the university's admission info's link
                link = cells[i + 2].find('a')['href']
                print(name, date, link)
                # Extend the name to the list stored in the global variable `school_names`
                school_names.append(name)
                yield {
                    'name': name,
                    'date': date,
                    'link': link
                }
            else:
                print('中国美术学院 is excluded')
        # Handle the TypeError exception that occurs when no link or date is provided
        except TypeError:
            print(f"No date, link or both of the school {name} provided")
            # Not yield a dictionary of the school's information
            pass


def write_homepage_form_to_excel(generator, fname):
    """
    Write the information to an excel form
    :param generator: a generator that generates a dictionary containing the school's name and its corresponding date\
     and url
    :type generator: generator
    :param fname: the name of the created excel file
    :type fname: str
    """
    sheet.title = '报考简章'
    sheet.cell(row=1, column=1).value = '高校名单'
    sheet.cell(row=1, column=2).value = '报名时间'
    sheet.cell(row=1, column=3).value = '招生简章'
    row = 2
    for school in generator:
        sheet.cell(row=row, column=1).value = school['name']
        sheet.cell(row=row, column=2).value = school['date']
        sheet.cell(row=row, column=3).value = school['link']
        row += 1

    wb.save(f'{fname}.xlsx')
    print(f"{fname}.xlsx saved")


def get_link(school_name):
    """
    Get the the admission guide link of the school with the passed name
    :param school_name: the name of the school
    :type school_name: str
    :return: the admission guide link
    :rtype: str
    """
    wb = openpyxl.load_workbook('浙江省2019年三位一体招生信息.xlsx')
    ws = wb.active
    row = school_names.index(school_name) + 2
    return ws.cell(row=row, column=3).value


def get_one_page_admission_guide(url):
    """
    Get the content of one page of the admission guide
    :param url: the url of the web page
    :type url: str
    :return: the content of on page of the admission guide
    :rtype: str
    """
    text = ""
    html = get_html(url)
    # Create a BeautifulSoup object using the parser html5lib
    soup = BeautifulSoup(html, features='html5lib')
    # Perform a CSS selection on the BeautifulSoup element to get all the cells of the form
    tags = soup.select('.TRS_Editor > p')
    for tag in tags:
        text = ''.join((text, tag.text, '\n'))
    if text:
        print(text)
        return text
    else:
        print("No data retrieved from the page")
        return None


def get_admission_guide(url):
    """
    Get the content of admission guide
    :param url: the url of the first page of the admission guide
    :type url: str
    :return: the content of admission guide
    :rtype: str
    """
    text = ''
    index = 0
    page_url = url
    pos = url.index('.shtml')
    # Get the html content of the first page
    page_1_text = get_one_page_admission_guide(page_url)
    # print(page_1_text)
    # print(type(page_1_text))
    text = "".join((text, page_1_text))
    while True:
        index += 1
        page_url = ''.join((url[:pos], '_', str(index), url[pos:]))
        page_text = get_one_page_admission_guide(page_url)
        # print(page_text)
        if page_text is None:
            print("Retrieve admission info guide. Done.")
            break
        else:
            text = ''.join((text, page_text))
    return text


def parse_admission_guide(html):
    pass


def write_admission_guide_to_excel():
    pass


if __name__ == '__main__':
    # homepage_html = get_homepage_html()
    # school_generator = parse_homepage(homepage_html)
    # write_homepage_form_to_excel(school_generator, "浙江省2019年三位一体招生信息")
    # admin_guide = get_admission_guide("http://gaokao.eol.cn/zhe_jiang/dongtai/201802/t20180211_1585545.shtml")
    # print(admin_guide)
    admin_guide = get_admission_guide("http://gaokao.eol.cn/zhe_jiang/dongtai/201802/t20180211_1585545.shtml")
    print(admin_guide)
